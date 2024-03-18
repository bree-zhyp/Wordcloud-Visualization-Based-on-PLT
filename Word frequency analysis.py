import pandas as pd  
from collections import Counter 
from openpyxl import Workbook, load_workbook
import os
import argparse  
import subprocess   

def run_get_wordcloud(args1):
    file = 'Get Wordcloud.py'
    subprocess.run(['python', file] + args1)

parser = argparse.ArgumentParser(description='Word frequency analysis')  
parser.add_argument('input_file', type=str, help='Path to the input text file')  
parser.add_argument('output_file', type=str, help='Path to the output Excel file')
parser.add_argument('wordcloud_file', type=str, help='Path to the output wordcloud file')  
args = parser.parse_args()  

def append_to_excel(data, excel_path):  
    wb = load_workbook(excel_path)  
    ws = wb['Sheet1']  

    ws.append(data)
    wb.save(excel_path)

# 读取Excel文件  
excel_path = args.input_file  
df = pd.read_excel(excel_path, usecols=['Word', 'POS'])  

# 读取stopword文件  
stopword_path = 'output/Text Annotation after Stop Word Filtering/stopwords.txt'  
stopword = set()

# 打开文件并读取内容  
with open(stopword_path, 'r', encoding='utf-8') as file:  
    for line in file:  
        # 移除每行两端的空白字符，包括换行符  
        word = line.strip() 
        # 将词添加到集合中  
        stopword.add(word)    

# 过滤出名词性词语  
nouns = df[df['POS'].str.startswith('n')]['Word']  
  
# 统计名词的词频  
noun_counts = Counter(nouns)  
  
# 打印词频统计结果
output_path = args.output_file  

# 初始化Excel文件并创建Sheet1(如果不存在)  
if not os.path.exists(output_path):  
    wb = Workbook()  
    ws = wb.active  
    ws.title = 'Sheet1'  
    ws.append(['noun', 'count'])  # 添加表头  
    wb.save(output_path) 

resource = []
datas = [(word, count) for word, count in noun_counts.most_common()]  # 确保只有word和count  
for data in datas:
    if(data[0] not in stopword):
        append_to_excel(data, output_path)
        if len(resource) < 30:
            resource.append(data[0])

print("Data after stop word filtering has been written to", output_path)

resource_string = ' '.join(resource)
wordcloud_path = args.wordcloud_file
arg = [resource_string, wordcloud_path]
run_get_wordcloud(arg)

