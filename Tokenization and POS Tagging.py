'''
Tokenization and part-of-speech tagging
这块Python代码主要用于对文本文件进行分词和词性标注,并将结果文件保存到output/Full Text Annotation文件夹中。
主要有以下几块内容:
  (1)导入库和模块;
  (2)初始化LTP工具;
  (3)分词、词性标注函数定义;
  (4)文件分块读取函数定义;
  (5)追加Excel文件写函数定义;
'''

from ltp import LTP  
import pandas as pd  
import os
from openpyxl import Workbook, load_workbook
import argparse  
  
# 定义命令行参数解析  
parser = argparse.ArgumentParser(description='Tokenization and part-of-speech tagging with LTP')  
parser.add_argument('input_file', type=str, help='Path to the input text file')  
parser.add_argument('output_file', type=str, help='Path to the output Excel file')  
args = parser.parse_args()  

ltp = LTP("small")  
  
def read_txt_file_chunks(file_path, chunk_size=16*32):  
    # 受限于LTP的最大读取字符数(500个),需要分块读入txt文件数据(chunk_size=16*32)
    with open(file_path, 'r', encoding='utf-8') as file:  
        while True:  
            chunk = file.read(chunk_size)  
            if not chunk:  
                break  
            yield chunk  
  
def process_chunk(chunk):  
    # 调用哈工大ltp进行分词、词性标注
    words = ltp.pipeline(chunk, tasks=["cws", "pos"])  
    return list(zip(words.cws, words.pos))  
  
def append_to_excel(data, excel_path):  
    # 加载现有的 Excel 文件  
    wb = load_workbook(excel_path)  
    ws = wb['Sheet1']  
      
    # 追加数据到工作表  
    for word, pos in data:  
        ws.append([word, pos])  
      
    # 保存工作簿  
    wb.save(excel_path)  
  
# file_path = 'dataset/十六大报告全文.txt'  
# excel_path = 'output/Full Text Annotation/FTA_十六大报告全文.xlsx'  
# 使用命令行参数替换硬编码的文件路径  
file_path = args.input_file  
excel_path = args.output_file 
  
# 初始化Excel文件并创建Sheet1(如果不存在)  
if not os.path.exists(excel_path):  
    wb = Workbook()  
    ws = wb.active  
    ws.title = 'Sheet1'  
    ws.append(['Word', 'POS'])  # 添加表头  
    wb.save(excel_path)  
  
# 使用生成器分块读取文件并处理  
for chunk in read_txt_file_chunks(file_path):  
    chunk_data = process_chunk(chunk)  
    append_to_excel(chunk_data, excel_path)  
  
print("Data has been written to", excel_path)